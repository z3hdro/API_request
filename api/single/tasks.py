from api.celery import app
from single.models import Link
from openpyxl import load_workbook
from datetime import datetime
from time import time
import requests

RESPONSE_TIMEOUT = 5


@app.task
def imports():
    """
    Функция - Celery task - запускает импорт данных из файла data.xlsx в БД.

    Возвращает строку с описанием статуса работы Celery task-a.
    """
    try:
        wb = load_workbook('./data.xlsx')
        ws = wb['Data']
    except FileNotFoundError:
        return 'File does not exist'
    except KeyError:
        return 'Sheet with name Data was not found!'
    else:
        for cell in ws['A']:
            try:
                link = Link.objects.get(link=cell.value)
            except Link.DoesNotExist:
                new_link = Link(link=cell.value)
                new_link.save()
        return 'links were imported successfully!'


@app.task
def check_link(url):
    """
    Функция - Celery task - получает URL-адрес,
    на который необходимо отправить запрос и получить ответ.
    Таймаут для запроса = 5 секундам. Если таймаут превышен,
    то в базу данных записывается Runtime error

    Возвращает описание работы Celery task-a

    :param url: URL-адрес
    :type url: string
    """
    link = Link.objects.get(link=url)
    start_time = datetime.now()
    current = time()
    try:
        response = requests.get('https://' + url, timeout=RESPONSE_TIMEOUT)
    except requests.exceptions.RequestException:
        link.status = -1
        link.description = 'Runtime error'
    else:
        link.status = response.status_code
        link.timeout = time() - current
        link.time = str(start_time)
        link.description = 'Success'
    finally:
        link.save()
        return 'Done'
